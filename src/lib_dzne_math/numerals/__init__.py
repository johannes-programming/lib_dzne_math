from openpyxl.utils import get_column_letter as _get_column_letter
from roman import toRoman as roman


def alpha(number):
    return _get_column_letter(number + 1)
